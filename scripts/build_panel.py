"""
build_panel.py
--------------
Builds a country×year panel dataset merging:
  - Oxford Insights Government AI Readiness Index (2021, 2022, 2023)
  - GitHub productivity metrics (2022, 2024 test runs)

Outputs: data/panel_dataset.csv
"""

import json
import pathlib
import re

import openpyxl
import pandas as pd

ROOT = pathlib.Path(__file__).parent.parent
DATA = ROOT / "data"
OXFORD_DIR = DATA / "oxford_insights"

# ---------------------------------------------------------------------------
# Country name → ISO 3166-1 alpha-2 mapping
# ---------------------------------------------------------------------------
COUNTRY_NAME_MAP = {
    "united states of america": "US",
    "united states": "US",
    "singapore": "SG",
    "united kingdom": "GB",
    "united kingdom of great britain and northern ireland": "GB",
    "finland": "FI",
    "canada": "CA",
    "france": "FR",
    "republic of korea": "KR",
    "south korea": "KR",
    "korea, republic of": "KR",
    "germany": "DE",
    "japan": "JP",
    "australia": "AU",
    "sweden": "SE",
    "netherlands": "NL",
    "denmark": "DK",
    "new zealand": "NZ",
    "norway": "NO",
    "austria": "AT",
    "switzerland": "CH",
    "israel": "IL",
    "china": "CN",
    "estonia": "EE",
    "ireland": "IE",
    "spain": "ES",
    "belgium": "BE",
    "luxembourg": "LU",
    "iceland": "IS",
    "portugal": "PT",
    "czech republic": "CZ",
    "czechia": "CZ",
    "italy": "IT",
    "malta": "MT",
    "taiwan": "TW",
    "taiwan, province of china": "TW",
    "russian federation": "RU",
    "russia": "RU",
    "brazil": "BR",
    "india": "IN",
    "ukraine": "UA",
    "bangladesh": "BD",
    "poland": "PL",
    "pakistan": "PK",
    "kenya": "KE",
    "egypt": "EG",
    "turkey": "TR",
    "türkiye": "TR",
    "latvia": "LV",
    "lithuania": "LT",
    "hungary": "HU",
    "croatia": "HR",
    "slovakia": "SK",
    "slovenia": "SI",
    "romania": "RO",
    "bulgaria": "BG",
    "greece": "GR",
    "cyprus": "CY",
    "united arab emirates": "AE",
    "saudi arabia": "SA",
    "qatar": "QA",
    "bahrain": "BH",
    "kuwait": "KW",
    "oman": "OM",
    "south africa": "ZA",
    "nigeria": "NG",
    "ghana": "GH",
    "morocco": "MA",
    "tunisia": "TN",
    "mexico": "MX",
    "argentina": "AR",
    "colombia": "CO",
    "chile": "CL",
    "peru": "PE",
    "uruguay": "UY",
    "costa rica": "CR",
    "panama": "PA",
    "malaysia": "MY",
    "thailand": "TH",
    "indonesia": "ID",
    "philippines": "PH",
    "vietnam": "VN",
    "sri lanka": "LK",
    "nepal": "NP",
    "myanmar": "MM",
    "cambodia": "KH",
    "kazakhstan": "KZ",
    "uzbekistan": "UZ",
    "azerbaijan": "AZ",
    "georgia": "GE",
    "armenia": "AM",
    "moldova": "MD",
    "belarus": "BY",
    "north macedonia": "MK",
    "serbia": "RS",
    "bosnia and herzegovina": "BA",
    "albania": "AL",
    "kosovo": "XK",
    "montenegro": "ME",
    "jordan": "JO",
    "lebanon": "LB",
    "iran": "IR",
    "iran, islamic republic of": "IR",
    "iraq": "IQ",
    "ethiopia": "ET",
    "tanzania": "TZ",
    "uganda": "UG",
    "rwanda": "RW",
    "senegal": "SN",
    "cameroon": "CM",
    "côte d'ivoire": "CI",
    "cote d'ivoire": "CI",
    "zambia": "ZM",
    "zimbabwe": "ZW",
    "mozambique": "MZ",
    "angola": "AO",
    "democratic republic of the congo": "CD",
    "congo, democratic republic of the": "CD",
    "venezuela": "VE",
    "venezuela, bolivarian republic of": "VE",
    "ecuador": "EC",
    "bolivia": "BO",
    "paraguay": "PY",
    "honduras": "HN",
    "guatemala": "GT",
    "el salvador": "SV",
    "nicaragua": "NI",
    "dominican republic": "DO",
    "cuba": "CU",
    "haiti": "HT",
    "jamaica": "JM",
    "trinidad and tobago": "TT",
}


def name_to_iso2(name: str) -> str | None:
    """Convert country name to ISO2 code. Returns None if not found."""
    key = name.strip().lower()
    # remove trailing punctuation or parenthetical
    key = re.sub(r"\s*\(.*\)", "", key).strip()
    return COUNTRY_NAME_MAP.get(key)


# ---------------------------------------------------------------------------
# Load Oxford Insights data
# ---------------------------------------------------------------------------
def load_oxford_2021() -> pd.DataFrame:
    wb = openpyxl.load_workbook(OXFORD_DIR / "2021-Government-AI-Readiness-Index.xlsx")
    ws = wb["Global ranking"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    records = []
    for row in rows:
        if row[0] is None:
            continue
        ranking, country_name, score = row[0], row[1], row[2]
        if not isinstance(ranking, (int, float)):
            continue
        iso2 = name_to_iso2(str(country_name))
        if iso2 and score is not None:
            records.append({"country": iso2, "year": 2021, "ai_readiness_score": float(score)})
    return pd.DataFrame(records)


def load_oxford_2022() -> pd.DataFrame:
    wb = openpyxl.load_workbook(OXFORD_DIR / "2022-Government-AI-Readiness-Index.xlsx")
    ws = wb["Global rankings"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # row 1 is header: ranking, country, total
        if row[0] is None:
            continue
        ranking, country_name, score = row[0], row[1], row[2]
        if not isinstance(ranking, (int, float)):
            continue
        if not isinstance(country_name, str):
            continue
        iso2 = name_to_iso2(country_name)
        if iso2 and score is not None:
            records.append({"country": iso2, "year": 2022, "ai_readiness_score": float(score)})
    return pd.DataFrame(records)


def load_oxford_2023() -> pd.DataFrame:
    wb = openpyxl.load_workbook(OXFORD_DIR / "2023-Government-AI-Readiness-Index.xlsx")
    ws = wb["Global rankings"]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # row 1 is header
        if row[0] is None:
            continue
        ranking, country_name, score = row[0], row[1], row[2]
        if not isinstance(ranking, (int, float)):
            continue
        iso2 = name_to_iso2(str(country_name))
        if iso2 and score is not None:
            records.append({"country": iso2, "year": 2023, "ai_readiness_score": float(score)})
    return pd.DataFrame(records)


def load_oxford_all() -> pd.DataFrame:
    dfs = [load_oxford_2021(), load_oxford_2022(), load_oxford_2023()]
    df = pd.concat(dfs, ignore_index=True)
    print(f"Oxford data: {len(df)} rows across years {sorted(df['year'].unique())}")
    print(f"  Countries with data: {df['country'].nunique()}")
    return df


# ---------------------------------------------------------------------------
# Load GitHub productivity data
# ---------------------------------------------------------------------------
def load_github_data() -> pd.DataFrame:
    with open(DATA / "github_productivity_results.json") as f:
        raw = json.load(f)

    records = []
    for year_str, year_data in raw.items():
        year = int(year_str)
        country_events = year_data.get("country_events", {})
        country_actors = year_data.get("country_actor_count", {})

        # collect all countries from either dict
        all_countries = set(country_events.keys()) | set(country_actors.keys())

        for country in all_countries:
            events = country_events.get(country, {})
            n_dev = country_actors.get(country, 0)

            commits = events.get("commits", 0)
            prs = events.get("pull_requests", 0)
            creates = events.get("creates", 0)
            comments = events.get("comments", 0)
            total = commits + prs + creates + comments

            if n_dev == 0:
                continue  # avoid division by zero

            records.append(
                {
                    "country": country,
                    "year": year,
                    "n_developers": n_dev,
                    "commits_per_dev": commits / n_dev,
                    "prs_per_dev": prs / n_dev,
                    "creates_per_dev": creates / n_dev,
                    "comments_per_dev": comments / n_dev,
                    "total_events_per_dev": total / n_dev,
                }
            )

    df = pd.DataFrame(records)
    print(f"GitHub data: {len(df)} rows across years {sorted(df['year'].unique())}")
    print(f"  Countries with data: {df['country'].nunique()}")
    return df


# ---------------------------------------------------------------------------
# Build panel
# ---------------------------------------------------------------------------
def build_panel() -> pd.DataFrame:
    oxford = load_oxford_all()
    github = load_github_data()

    # GitHub years are 2022 and 2024; Oxford data is 2021, 2022, 2023
    # For 2022 GitHub data: use Oxford 2022 AI readiness score
    # For 2024 GitHub data: use Oxford 2023 AI readiness score (latest available)
    oxford_lookup = oxford.set_index(["country", "year"])["ai_readiness_score"].to_dict()

    def get_ai_score(country, github_year):
        if github_year == 2022:
            return oxford_lookup.get((country, 2022))
        elif github_year == 2024:
            return oxford_lookup.get((country, 2023))
        return None

    github["ai_readiness_score"] = github.apply(
        lambda r: get_ai_score(r["country"], r["year"]), axis=1
    )

    panel = github[github["ai_readiness_score"].notna()].copy()
    panel = panel[
        ["country", "year", "ai_readiness_score", "commits_per_dev", "prs_per_dev",
         "creates_per_dev", "comments_per_dev", "total_events_per_dev", "n_developers"]
    ].sort_values(["country", "year"]).reset_index(drop=True)

    print(f"\nPanel dataset: {len(panel)} rows")
    print(f"  Countries: {sorted(panel['country'].unique())}")
    print(f"  Years: {sorted(panel['year'].unique())}")
    print(panel.to_string())

    out_path = DATA / "panel_dataset.csv"
    panel.to_csv(out_path, index=False)
    print(f"\nSaved to {out_path}")
    return panel


if __name__ == "__main__":
    build_panel()
